import time
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook

# =================【Google Sheet 連線】=================
def load_sheet(sheet_url, sheet_name):
    scope = ["https://spreadsheets.google.com/feeds",
             "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name("service_account.json", scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_url(sheet_url).worksheet(sheet_name)
    return sheet

# =================【初始化 Selenium】=================
def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    return driver

# =================【登入非登不可】=================
def login_fadenbook(driver, company_id, pin_code, id_number):
    driver.get("https://fadenbook.fda.gov.tw/pub/index.aspx")
    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.ID, "btnLogin"))
    ).click()
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "txt_Pin"))
    ).send_keys(pin_code)
    driver.find_element(By.ID, "MakeSignature").click()
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='text']"))
    ).send_keys(id_number)
    driver.find_element(By.CSS_SELECTOR, "input[type='submit']").click()

# =================【更新保險資料】=================
def update_insurance(driver, file_path, policy_no, start_date, end_date):
    upload_input = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "fileUpload"))
    )
    upload_input.send_keys(file_path)
    driver.find_element(By.ID, "btnUploadFile").click()

    driver.find_element(By.ID, "txt_InsureNo").clear()
    driver.find_element(By.ID, "txt_InsureNo").send_keys(policy_no)

    driver.find_element(By.ID, "txt_InsureBeginDate").clear()
    driver.find_element(By.ID, "txt_InsureBeginDate").send_keys(start_date)
    driver.find_element(By.ID, "txt_InsureEndDate").clear()
    driver.find_element(By.ID, "txt_InsureEndDate").send_keys(end_date)

    driver.find_element(By.ID, "btn_CompanyDrBusin_Save").click()
    time.sleep(3)

# =================【登出】=================
def logout(driver):
    driver.find_element(By.ID, "UMenu_ImgBut1").click()
    driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_btnOK").click()

# =================【批次流程】=================
def batch_process(sheet, driver, pin_code, id_number):
    rows = sheet.get_all_records()
    results = []

    for row in rows:
        company_id = row["統編"]
        file_path = row["保險檔案路徑"]
        policy_no = row["保單號碼"]
        start_date = row["保險起始"]
        end_date = row["保險到期"]

        try:
            login_fadenbook(driver, company_id, pin_code, id_number)
            update_insurance(driver, file_path, policy_no, start_date, end_date)
            logout(driver)
            results.append((company_id, "成功", ""))
        except Exception as e:
            results.append((company_id, "失敗", str(e)))

    return results

# =================【報表輸出】=================
def export_results(results, filename="upload_results.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "上傳結果"

    ws.append(["統編", "狀態", "原因"])
    for company_id, status, reason in results:
        ws.append([company_id, status, reason])

    wb.save(filename)
    print(f"📊 已輸出報表：{filename}")

# =================【讀取失敗統編】=================
def load_failed_companies(filename="upload_results.xlsx"):
    wb = load_workbook(filename)
    ws = wb.active
    failed_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳過標題列
        company_id, status, reason = row
        if status == "失敗":
            failed_list.append(company_id)

    return failed_list

# =================【失敗重跑流程】=================
def retry_failed(sheet, driver, pin_code, id_number, failed_list):
    rows = sheet.get_all_records()
    results = []

    for row in rows:
        company_id = row["統編"]
        if company_id not in failed_list:
            continue  # 只跑失敗的統編

        file_path = row["保險檔案路徑"]
        policy_no = row["保單號碼"]
        start_date = row["保險起始"]
        end_date = row["保險到期"]

        try:
            login_fadenbook(driver, company_id, pin_code, id_number)
            update_insurance(driver, file_path, policy_no, start_date, end_date)
            logout(driver)
            results.append((company_id, "成功", ""))
        except Exception as e:
            results.append((company_id, "失敗", str(e)))

    return results

# =================【主程式】=================
if __name__ == "__main__":
    sheet_url = "https://docs.google.com/spreadsheets/d/1HxNKWMyeXF1vG-QfMidbjmb2GmZIpi3P/edit#gid=717839605"
    sheet = load_sheet(sheet_url, "工作表1")

    driver = init_driver()
    pin_code = "你的憑證PIN"
    id_number = "你的身分證字號"

    # 第一次批次上傳
    results = batch_process(sheet, driver, pin_code, id_number)
    export_results(results, "upload_results.xlsx")

    # 讀取失敗統編並重跑
    failed_list = load_failed_companies("upload_results.xlsx")
    if failed_list:
        print(f"⚠️ 發現失敗統編：{failed_list}")
        retry_results = retry_failed(sheet, driver, pin_code, id_number, failed_list)
        export_results(retry_results, "retry_results.xlsx")

    driver.quit()
